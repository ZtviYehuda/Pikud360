"""
Reports module Excel processor.
Generates structured Excel workbooks using openpyxl.
"""
import io
from datetime import datetime, date
from typing import Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.modules.reports.models import ReportRequest
from app.modules.reports.processors.base import ReportProcessor

class ExcelReportProcessor(ReportProcessor):
    """Processes Excel workbook report generation using openpyxl."""

    def validate(self, report: ReportRequest) -> None:
        # Check standard properties if required
        pass

    def generate(self, report: ReportRequest) -> bytes:
        headers, rows = self._fetch_data(report)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Report Data"
        
        # 1. Write header row
        ws.append(headers)
        
        # 2. Style headers (bold)
        header_font = Font(bold=True)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            
        # 3. Write data rows & format numeric / date fields
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                
                # Check data type
                if isinstance(val, (datetime, date)):
                    cell.value = val.isoformat()
                else:
                    cell.value = val
                
                # Implicit Number Formatting
                if isinstance(val, int):
                    cell.number_format = '#,##0'
                elif isinstance(val, float):
                    cell.number_format = '#,##0.00'
                    
        # 4. Freeze top row (headers remain visible on scroll)
        ws.freeze_panes = "A2"
        
        # 5. Enable Auto-Filters
        last_col_letter = get_column_letter(len(headers))
        # Wait, if there are no rows, set filter only on the header row itself
        max_row = max(len(rows) + 1, 1)
        ws.auto_filter.ref = f"A1:{last_col_letter}{max_row}"
        
        # 6. Adjust column widths automatically based on max cell length
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            # Add padding
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        # 7. Write to BytesIO buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        
        return excel_bytes
